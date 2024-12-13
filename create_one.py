import pandas as pd
import os
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Define color fills for Excel
color_fills = {
    "red": PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),
    "green": PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid"),
    "orange": PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid"),
    "yellow": PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"),
    "white": PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
}

# Define service categories
categories = {
    'Security and Identity': ['IAM', 'ACM', 'KMS', 'GuardDuty', 'Secret Manager', 'Secret Hub', 'SSM'],
    'Compute': ['Auto Scaling', 'EC2', 'ECS', 'EKS', 'Lambda', 'EMR', 'Step Functions'],
    'Storage': ['EBS', 'ECR', 'S3', 'DLM', 'Backup'],
    'Network': ['API Gateway', 'CloudFront', 'Route 53', 'VPC', 'ELB', 'ElasticCache', 'CloudTrail'],
    'Database': ['RDS', 'DynamoDB', 'Athena', 'Glue'],
    'Other': ['CloudFormation', 'CodeDeploy', 'Config', 'SNS', 'SQS', 'WorkSpaces', 'EventBridge', 'Config']
}

def load_input_file(input_file):
    """
    Load input file (CSV or Excel) with error handling
    """
    try:
        if input_file.endswith(".csv"):
            return pd.read_csv(input_file, low_memory=False)
        elif input_file.endswith((".xlsx", ".xls")):
            return pd.read_excel(input_file, engine='openpyxl')
        else:
            raise ValueError("Unsupported file type. Please use CSV or Excel files.")
    except Exception as e:
        print(f"Error loading input file: {e}")
        raise

def load_priority_database(priority_file="PowerPipeControls_Annotations.xlsx"):
    """
    Load priority database with error handling
    """
    try:
        return pd.read_excel(priority_file)
    except Exception as e:
        print(f"Error loading priority database: {e}")
        raise

def update_priority_and_recommendation(df_input, df_priority):
    """
    Update input DataFrame with priority and recommendations from database
    """
    for idx, row in df_input.iterrows():
        control_title = row.get("control_title", "")
        status = row.get("status", "")

        # Search for the control_title in the priority database
        matching_row = df_priority[df_priority["control_title"] == control_title]

        if not matching_row.empty:
            priority = matching_row.iloc[0]["priority"]
            recommendation = matching_row.iloc[0]["Recommendation Steps/Approach"]

            # Assign priority and recommendation
            if status in ["ok", "info", "skip"]:
                df_input.at[idx, "priority"] = "Safe/Well Architected"
                df_input.at[idx, "Recommendation Steps/Approach"] = recommendation
                df_input.at[idx, "priority_color"] = "green"
            else:
                df_input.at[idx, "priority"] = priority
                df_input.at[idx, "Recommendation Steps/Approach"] = recommendation
                
                # Assign color based on priority
                if priority == "High":
                    df_input.at[idx, "priority_color"] = "red"
                elif priority == "Medium":
                    df_input.at[idx, "priority_color"] = "orange"
                elif priority == "Low":
                    df_input.at[idx, "priority_color"] = "yellow"
        else:
            # If no match is found, set default values
            df_input.at[idx, "priority"] = "No data"
            df_input.at[idx, "Recommendation Steps/Approach"] = "No recommendation available"
            df_input.at[idx, "priority_color"] = "white"

    return df_input

def create_simplified_report(df_input, final_report_file):
    """
    Create a simplified report with categorized sheets
    """
    # Add 'fixed' and 'feedback' columns (only for category sheets)
    df_input['fixed'] = ''
    df_input['feedback'] = ""

    # Filter data based on status
    unsafe_df = df_input[df_input['status'] == 'alarm'].copy()
    safe_df = df_input[df_input['status'] != 'alarm'].copy()

    # Remove 'fixed' and 'feedback' columns from "safe" and "unsafe"
    safe_df = safe_df.drop(columns=['fixed', 'feedback'])
    unsafe_df = unsafe_df.drop(columns=['fixed', 'feedback'])

    # Initialize a dictionary to hold DataFrames for each category
    categorized_data = {category: pd.DataFrame() for category in categories}

    # Process each category and filter relevant services
    for category, services in categories.items():
        categorized_data[category] = unsafe_df[unsafe_df['title'].isin(services)]

    return safe_df, unsafe_df, categorized_data

def write_output_file(safe_df, unsafe_df, categorized_data, final_report_file, df_input):
    """
    Write output file with multiple sheets and formatting
    """
    with pd.ExcelWriter(final_report_file, engine='xlsxwriter') as writer:
        workbook = writer.book

        # Define formats
        header_format = workbook.add_format({'bold': True, 'bg_color': '#FFA07A', 'font_color': 'black'})
        high_priority_format = workbook.add_format({'bg_color': '#FF0000', 'font_color': 'white'})
        medium_priority_format = workbook.add_format({'bg_color': '#FFA500', 'font_color': 'black'})
        low_priority_format = workbook.add_format({'bg_color': '#FFFF00', 'font_color': 'black'})
        safe_priority_format = workbook.add_format({'bg_color': '#008000', 'font_color': 'white'})

        # Write raw data sheet first
        df_input.to_excel(writer, sheet_name='Raw Data', index=False)
        worksheet = writer.sheets['Raw Data']
        for col_num, value in enumerate(df_input.columns.values):
            worksheet.write(0, col_num, value, header_format)

        # Write "safe" and "unsafe" DataFrames
        for sheet_name, df_data in [('Safe', safe_df), ('Unsafe', unsafe_df)]:
            df_data.to_excel(writer, sheet_name=sheet_name, index=False)
            worksheet = writer.sheets[sheet_name]
            for col_num, value in enumerate(df_data.columns.values):
                worksheet.write(0, col_num, value, header_format)

        # Write each category DataFrame
        for category, data in categorized_data.items():
            if not data.empty:
                data.to_excel(writer, sheet_name=category, index=False)
                worksheet = writer.sheets[category]
                for col_num, value in enumerate(data.columns.values):
                    worksheet.write(0, col_num, value, header_format)

    print(f"Final simplified report saved as {final_report_file}")

def main():
    # Ask the user to input the report file name
    input_file = input("Enter the input file name (CSV or Excel): ").strip()

    # Get the input file's base name and create unique output file name
    base_name = os.path.splitext(os.path.basename(input_file))[0]
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    final_report_file = f"{base_name}_comprehensive_report_{timestamp}.xlsx"

    try:
        # Load input file and priority database
        df_input = load_input_file(input_file)
        df_priority = load_priority_database()

        # Update priority and recommendations
        updated_df = update_priority_and_recommendation(df_input, df_priority)

        # Create simplified report
        safe_df, unsafe_df, categorized_data = create_simplified_report(updated_df, final_report_file)

        # Write output file
        write_output_file(safe_df, unsafe_df, categorized_data, final_report_file, updated_df)

    except Exception as e:
        print(f"An error occurred: {e}")

if __name__ == "__main__":
    main()

###the output from this above program get used as input in this below program
import pandas as pd
import os
from datetime import datetime
import xlsxwriter

# Define service categories
categories = {
    'Security and Identity': ['IAM', 'ACM', 'KMS', 'GuardDuty', 'Secret Manager', 'Secret Hub', 'SSM'],
    'Compute': ['Auto Scaling', 'EC2', 'ECS', 'EKS', 'Lambda', 'EMR', 'Step Functions'],
    'Storage': ['EBS', 'ECR', 'S3', 'DLM', 'Backup'],
    'Network': ['API Gateway', 'CloudFront', 'Route 53', 'VPC', 'ELB', 'ElasticCache', 'CloudTrail'],
    'Database': ['RDS', 'DynamoDB', 'Athena', 'Glue'],
    'Other': ['CloudFormation', 'CodeDeploy', 'Config', 'SNS', 'SQS', 'WorkSpaces', 'EventBridge', 'Config']
}

# Map numerical priorities to words
priority_map = {1: "High", 2: "Medium", 3: "Low"}

def create_simplified_report_with_pivot(report_file, final_report_file):
    # Read input report file (CSV or Excel)
    if report_file.endswith('.csv'):
        df = pd.read_csv(report_file)
    elif report_file.endswith(('.xls', '.xlsx')):
        df = pd.read_excel(report_file)
    else:
        raise ValueError("Unsupported file format. Please provide a CSV or Excel file.")
    
    # Ensure required columns exist
    required_columns = [
        'group_id', 'title', 'description', 'control_id', 'control_title', 
        'control_description', 'resource', 'status', 'priority'
    ]
    missing_columns = [col for col in required_columns if col not in df.columns]
    if missing_columns:
        raise KeyError(f"Missing columns: {', '.join(missing_columns)}")

    # Clean and prepare data
    df['status'] = df['status'].astype(str)
    df['is_open_issue'] = df['status'].apply(lambda x: 1 if x == 'alarm' else 0)
    df['priority'] = df['priority'].map(priority_map).fillna(df['priority'])

    # Add new columns for analysis
    df['fixed'] = False  # Checkbox column
    df['feedback'] = ''  # Feedback column

    # Prepare columns for different sheets
    raw_data_columns = [
        'group_id', 'title', 'description', 'control_id', 'control_title', 
        'resource', 'status', 'priority', 'Recommendation Steps/Approach', 
        'fixed', 'feedback'
    ]

    # Ensure all required columns exist in the raw data
    raw_data_df = df[raw_data_columns]

    # Filter rows for different analysis
    no_open_issues_df = df[df['is_open_issue'] == 0]
    open_issues_df = df[df['is_open_issue'] >= 1]

    # Create Excel writer
    with pd.ExcelWriter(final_report_file, engine='xlsxwriter') as writer:
        # Write Raw Data Sheet
        raw_data_df.to_excel(writer, sheet_name='Raw Data', index=False)

        # Write No Open Issues and Open Issues Sheets
        no_open_issues_df.to_excel(writer, sheet_name='no_open_issues', index=False)
        open_issues_df.to_excel(writer, sheet_name='open_issues', index=False)

        # Create Summary Data
        summary_data = []
        sr_no = 1

        # Process each service category
        for service in categories.keys():
            service_df = open_issues_df[open_issues_df['title'].isin(categories[service])]
            
            # Group by Control Title and sum Open Issues
            service_grouped = service_df.groupby(
                ['title', 'control_title', 'control_description', 'priority'], 
                as_index=False
            ).agg({'is_open_issue': 'sum'})

            # Add service heading
            summary_data.append({
                'Sr No': f'{service} Heading', 
                'Service': '', 
                'Control Title': '', 
                'Description': '', 
                'Open Issues': '', 
                'Priority': ''
            })

            # Add grouped data to summary
            for _, row in service_grouped.iterrows():
                summary_data.append({
                    'Sr No': sr_no,
                    'Service': row['title'],
                    'Control Title': row['control_title'],
                    'Description': row['control_description'],
                    'Open Issues': row['is_open_issue'],
                    'Priority': row['priority']
                })
                sr_no += 1

            # Add blank row between services
            summary_data.append({
                'Sr No': '',
                'Service': '',
                'Control Title': '',
                'Description': '',
                'Open Issues': '',
                'Priority': ''
            })

        # Convert summary to DataFrame
        summary_df = pd.DataFrame(summary_data)

        # Create custom summary sheet with color formatting
        workbook = writer.book
        
        # Define color formats
        red_format = workbook.add_format({'bg_color': 'red', 'font_color': 'white'})
        orange_format = workbook.add_format({'bg_color': 'orange', 'font_color': 'white'})
        yellow_format = workbook.add_format({'bg_color': 'yellow', 'font_color': 'black'})
        green_format = workbook.add_format({'bg_color': 'lightgreen', 'font_color': 'black'})

        # Write detailed summary sheet with priority colors
        summary_df.to_excel(writer, sheet_name='detailed_summary', index=False)
        worksheet = writer.sheets['detailed_summary']

        # Apply color formatting
        for row_num, row in enumerate(summary_df.values, start=1):
            priority = row[5]
            if priority == 'High':
                worksheet.write(row_num, 5, row[5], red_format)
            elif priority == 'Medium':
                worksheet.write(row_num, 5, row[5], orange_format)
            elif priority == 'Low':
                worksheet.write(row_num, 5, row[5], yellow_format)

        # Pivot Tables and Analysis
        # Open Issues Analysis
        analysis_df = pd.pivot_table(
            open_issues_df, 
            values='is_open_issue', 
            index=['title'], 
            aggfunc='sum', 
            fill_value=0
        )
        analysis_df.to_excel(writer, sheet_name='open_issues_analysis')

        # Safe Controls Analysis
        safe_controls_pivot = pd.pivot_table(
            no_open_issues_df, 
            values='is_open_issue', 
            index=['title'], 
            aggfunc='count', 
            fill_value=0
        )
        safe_controls_pivot.to_excel(writer, sheet_name='safe_controls_analysis')

        # Charting
        worksheet_open_analysis = writer.sheets['open_issues_analysis']
        worksheet_safe_analysis = writer.sheets['safe_controls_analysis']

        # Chart for Open Issues
        chart1 = workbook.add_chart({'type': 'column'})
        chart1.add_series({
            'name': 'Open Issues',
            'categories': f'=open_issues_analysis!$A$2:$A${len(analysis_df)+1}',
            'values': f'=open_issues_analysis!$B$2:$B${len(analysis_df)+1}'
        })
        worksheet_open_analysis.insert_chart('D2', chart1)

        # Chart for Safe Controls
        chart2 = workbook.add_chart({'type': 'column'})
        chart2.add_series({
            'name': 'Safe Controls (No Issues)',
            'categories': f'=safe_controls_analysis!$A$2:$A${len(safe_controls_pivot)+1}',
            'values': f'=safe_controls_analysis!$B$2:$B${len(safe_controls_pivot)+1}'
        })
        worksheet_safe_analysis.insert_chart('D2', chart2)

        # Priority Level Summary
        priority_summary = df['priority'].value_counts().to_dict()
        high_count = priority_summary.get("High", 0)
        medium_count = priority_summary.get("Medium", 0)
        low_count = priority_summary.get("Low", 0)
        blank_count = len(df[df['priority'] == 'No Priority'])
        grand_total = high_count + medium_count + low_count + blank_count

        # Add summary to a dedicated sheet
        summary_sheet = workbook.add_worksheet('Priority Summary')
        summary_sheet.write('A1', "Priority Level Summary")
        summary_sheet.write('A3', f"High Priority: {high_count}")
        summary_sheet.write('A4', f"Medium Priority: {medium_count}")
        summary_sheet.write('A5', f"Low Priority: {low_count}")
        summary_sheet.write('A6', f"No Priority: {blank_count}")
        summary_sheet.write('A7', f"Grand Total: {grand_total}")
    
  # Count priorities
    priority_counts = df['priority'].value_counts().reindex(
        ['High', 'Medium', 'Low', 'Safe/Well Architected', 'No Priority'], fill_value=0
    )

    # Convert counts to a DataFrame for the summary
    summary_df = priority_counts.reset_index()
    summary_df.columns = ['Priority', 'Count']

    # Add Total Row
    total_row = pd.DataFrame([['Total', summary_df['Count'].sum()]], columns=['Priority', 'Count'])
    summary_df = pd.concat([summary_df, total_row], ignore_index=True)

    # Create Excel writer
    with pd.ExcelWriter(final_report_file, engine='xlsxwriter') as writer:
        workbook = writer.book

        # Write Priority Summary Sheet
        summary_df.to_excel(writer, sheet_name='Priority Summary', index=False)
        worksheet = writer.sheets['Priority Summary']

        # Define chart and color mappings
        chart = workbook.add_chart({'type': 'column'})
        color_map = {
            'High': '#FF0000',  # Red
            'Medium': '#FFA500',  # Orange
            'Low': '#FFFF00',  # Yellow
            'Safe/Well Architected': '#00FF00',  # Green
            'No Priority': '#C0C0C0'  # Gray
        }

        # Add data to chart
        for row_num, row in summary_df.iterrows():
            priority = row['Priority']
            if priority != 'Total':
                chart.add_series({
                    'name': priority,
                    'categories': f"='Priority Summary'!$A$2:$A${len(summary_df)}",
                    'values': f"='Priority Summary'!$B${row_num + 2}:$B${row_num + 2}",
                    'fill': {'color': color_map.get(priority, '#000000')},
                })

        # Customize chart appearance
        chart.set_title({'name': 'Priority Analysis'})
        chart.set_x_axis({'name': 'Priority'})
        chart.set_y_axis({'name': 'Count'})
        chart.set_legend({'position': 'bottom'})

        # Insert chart into worksheet
        worksheet.insert_chart('D2', chart)
      
    print(f"Final report with pivot table and charts saved as {final_report_file}")

def main():
    # Ask the user to input the report file name
    report_file = input("Enter the report file name (e.g., aws_compliance_benchmark_all_controls_benchmark_vested_with_priorities.csv): ").strip()
    
    # Get the input file's base name (without path) and extension
    base_name = os.path.splitext(os.path.basename(report_file))[0]
    
    # Create a unique file name using a timestamp
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    unique_file_name = f"{base_name}_comprehensive_report_{timestamp}.xlsx"
    
    # Set the output path to the reports directory
    reports_directory = os.path.dirname(os.path.abspath(__file__))  # Get the current script's directory
    final_report_file = os.path.join(reports_directory, unique_file_name)

    # Ask if the user wants to create the simplified report
    create_report = input("Do you want to create the comprehensive report? (yes/no): ").strip().lower()
    if create_report == 'yes':
        try:
            create_simplified_report_with_pivot(report_file, final_report_file)
        except ValueError as e:
            print(e)
        except KeyError as e:
            print(f"KeyError: {e}")
        except Exception as e:
            print(f"An error occurred: {e}")
    else:
        print("Report creation skipped.")

if __name__ == "__main__":
    main()
