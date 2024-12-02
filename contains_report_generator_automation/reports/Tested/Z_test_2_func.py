import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Define color fills for Excel at the top level
green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# Load the input file and the database
def load_data(input_file, priority_file):
    if input_file.endswith(".xlsx"):
        df_input = pd.read_excel(input_file)
    elif input_file.endswith(".csv"):
        df_input = pd.read_csv(input_file)
    else:
        raise ValueError("Unsupported file type")

    # Load priority database
    df_priority = pd.read_excel(priority_file)
    
    return df_input, df_priority

# Match control_title and update with priority and recommendations
def update_priority_and_recommendation(df_input, df_priority):
    # Iterate over each row in the input data
    for idx, row in df_input.iterrows():
        control_title = row["control_title"]
        status = row["status"]

        # Search for the control_title in the priority database
        matching_row = df_priority[df_priority["control_title"] == control_title]

        if not matching_row.empty:
            priority = matching_row.iloc[0]["priority"]
            recommendation = matching_row.iloc[0]["Recommendation Steps/Approach"]

            # If the status is ok, info, or skip, color the priority green and set text
            if status in ["ok", "info", "skip"]:
                df_input.at[idx, "priority"] = "Safe/Well Architected"  # Specific value for "ok", "info", "skip"
                df_input.at[idx, "recommendation"] = ""
                df_input.at[idx, "priority_label"] = "Safe/Well Architected"
                df_input.at[idx, "priority_color"] = "00FF00"  # Green color

            else:
                # For other statuses, assign the appropriate priority and recommendation
                df_input.at[idx, "priority"] = priority
                df_input.at[idx, "recommendation"] = recommendation
                df_input.at[idx, "priority_label"] = priority
                
                # Color the priority column based on the priority
                if priority == "High":
                    df_input.at[idx, "priority_color"] = "FF0000"
                elif priority == "Medium":
                    df_input.at[idx, "priority_color"] = "FFA500"
                elif priority == "Low":
                    df_input.at[idx, "priority_color"] = "FFFF00"

        else:
            # If no match is found, set default values
            df_input.at[idx, "priority"] = "No data"
            df_input.at[idx, "recommendation"] = "No recommendation available"
            df_input.at[idx, "priority_label"] = "No data"
            df_input.at[idx, "priority_color"] = "FFFFFF"  # White or default

    return df_input

# Write the output file
def write_output(df_input, output_file):
    # Save the updated data frame to Excel
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df_input.to_excel(writer, index=False, sheet_name="Sheet1")
    
    # Now apply the color formatting to the priority column based on the 'priority_color' field
    wb = writer.book
    ws = wb["Sheet1"]
    for idx, row in df_input.iterrows():
        priority_color = row["priority_color"]
        if priority_color != "FFFFFF":  # Skip if no color is needed
            cell = ws.cell(row=idx+2, column=df_input.columns.get_loc("priority")+1)
            if priority_color == "FF0000":
                cell.fill = red_fill
            elif priority_color == "FFA500":
                cell.fill = orange_fill
            elif priority_color == "FFFF00":
                cell.fill = yellow_fill
            elif priority_color == "00FF00":
                cell.fill = green_fill

    wb.save(output_file)

def main():
    # Get file names from user
    input_file = input("Enter the input file name (CSV or Excel): ")
    priority_file = "PowerPipeControls_Annotations.xlsx"  # Database file containing control titles, priorities, etc.
    output_file = input_file.split(".")[0] + "_updated.xlsx"  # Save as Excel file with '_updated' suffix

    # Load data
    df_input, df_priority = load_data(input_file, priority_file)

    # Update input with priority and recommendations
    updated_df = update_priority_and_recommendation(df_input, df_priority)

    # Write output file with priority color formatting
    write_output(updated_df, output_file)
    print(f"Updated file saved as {output_file}")

if __name__ == "__main__":
    main()
